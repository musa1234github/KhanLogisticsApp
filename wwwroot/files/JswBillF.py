# JswBillF_multi_consolidate.py
"""
Consolidate multiple 'JSW BILL' Excel files into a single formatted sheet.

Features:
- Reads all .xlsx/.xls files in the folder (or given folder path).
- For each file:
  - Scans the top area (no header) to extract Bill No and Bill Date as plain text (no parsing).
  - Detects the table header row (looks for "Supply inv.Date", flexible).
  - Reads the table, maps columns to the required output schema.
  - Drops summary/total rows (where ChallanNo is non-numeric).
  - Appends Bill Num and BILL DATE (plain text) to each row.
- Concatenates all files' rows one after the other (no blank rows between files).
- Converts numeric-like columns (Dispatch Quantity, Rate, Total Price) to numeric types where possible
  so Excel SUM across the consolidated sheet will work.
- Writes a single Excel workbook "JSW_BILL_FORMATTED.xlsx" with a single sheet "JSW".
- Keeps header formatting and cell styling similar to previous code.

Usage:
- Just run:
    python JswBillF_multi_consolidate.py
"""

import os
import sys
import re
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, date

# ------------------------
# Utilities
# ------------------------
def _is_nonempty(cell):
    if cell is None:
        return False
    try:
        if isinstance(cell, float) and math.isnan(cell):
            return False
    except Exception:
        pass
    s = str(cell).strip()
    return s != "" and s.lower() != "nan"

def _is_meaningful_text(cell_text):
    if cell_text is None:
        return False
    s = str(cell_text).strip()
    if s == "":
        return False
    return re.search(r"[A-Za-z0-9]", s) is not None

def _to_text(cell):
    if cell is None:
        return ""
    try:
        import pandas as pd
        if isinstance(cell, (pd.Timestamp, datetime, date)):
            try:
                return cell.strftime("%d-%b-%y")
            except Exception:
                return str(cell)
    except Exception:
        pass
    return str(cell).strip()

def _first_neighbor_value(raw_df, r, c, max_right=8):
    cols = raw_df.shape[1]
    for dc in range(1, max_right + 1):
        cc = c + dc
        if cc < cols:
            v = raw_df.iat[r, cc]
            if _is_nonempty(v) and _is_meaningful_text(v):
                return _to_text(v)
    if r + 1 < raw_df.shape[0]:
        v = raw_df.iat[r + 1, c]
        if _is_nonempty(v) and _is_meaningful_text(v):
            return _to_text(v)
    if r + 1 < raw_df.shape[0] and c + 1 < cols:
        v = raw_df.iat[r + 1, c + 1]
        if _is_nonempty(v) and _is_meaningful_text(v):
            return _to_text(v)
    # try a couple more offsets below-right
    if r + 1 < raw_df.shape[0]:
        for dc in (2, 3):
            cc = c + dc
            if cc < cols:
                v = raw_df.iat[r + 1, cc]
                if _is_nonempty(v) and _is_meaningful_text(v):
                    return _to_text(v)
    return ""

def _fallback_scan_for_billnum(raw_df):
    max_rows = min(60, raw_df.shape[0])
    for r in range(max_rows):
        for c in range(raw_df.shape[1]):
            v = raw_df.iat[r, c]
            if not _is_nonempty(v):
                continue
            s = str(v).strip()
            if '/' in s and len(s) >= 4 and _is_meaningful_text(s):
                return _to_text(s)
    for r in range(max_rows):
        for c in range(raw_df.shape[1]):
            v = raw_df.iat[r, c]
            if not _is_nonempty(v):
                continue
            s = str(v).strip()
            if _is_meaningful_text(s) and len(s) >= 4:
                return _to_text(s)
    return ""

# ------------------------
# Extract metadata (Bill No, Bill Date)
# ------------------------
def _extract_bill_meta(raw_df):
    bill_num = ""
    bill_date = ""
    max_rows_to_scan = min(80, raw_df.shape[0])
    pat_no = re.compile(r"\bbill\s*no\b", re.IGNORECASE)
    pat_date = re.compile(r"\bbill\s*date\b", re.IGNORECASE)

    for r in range(max_rows_to_scan):
        for c in range(raw_df.shape[1]):
            cell = raw_df.iat[r, c]
            if not _is_nonempty(cell):
                continue
            text = str(cell)

            if pat_no.search(text) and not bill_num:
                m = re.search(r"bill\s*no\s*[:\-\–\—]?\s*(.+)$", text, flags=re.IGNORECASE)
                if m:
                    candidate = m.group(1).strip()
                    if _is_meaningful_text(candidate):
                        bill_num = _to_text(candidate)
                    else:
                        nb = _first_neighbor_value(raw_df, r, c)
                        if _is_meaningful_text(nb):
                            bill_num = nb
                else:
                    nb = _first_neighbor_value(raw_df, r, c)
                    if _is_meaningful_text(nb):
                        bill_num = nb

            if pat_date.search(text) and not bill_date:
                m = re.search(r"bill\s*date\s*[:\-\–\—]?\s*(.+)$", text, flags=re.IGNORECASE)
                if m:
                    candidate = m.group(1).strip()
                    if _is_meaningful_text(candidate):
                        bill_date = _to_text(candidate)
                    else:
                        nb = _first_neighbor_value(raw_df, r, c)
                        if _is_meaningful_text(nb):
                            bill_date = nb
                else:
                    nb = _first_neighbor_value(raw_df, r, c)
                    if _is_meaningful_text(nb):
                        bill_date = nb

            if bill_num and bill_date:
                return bill_num, bill_date

    if not bill_num:
        candidate = _fallback_scan_for_billnum(raw_df)
        if _is_meaningful_text(candidate):
            bill_num = candidate

    if not bill_date:
        # try find any date-like text in top area
        date_re = re.compile(r"\b(\d{1,2}[-/]\w{3,9}[-/]\d{2,4}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b")
        date_like = ""
        for r in range(min(80, raw_df.shape[0])):
            for c in range(raw_df.shape[1]):
                v = raw_df.iat[r, c]
                if not _is_nonempty(v):
                    continue
                s = str(v).strip()
                m = date_re.search(s)
                if m:
                    date_like = m.group(0).strip()
                    break
            if date_like:
                break
        if _is_meaningful_text(date_like):
            bill_date = _to_text(date_like)

    return bill_num, bill_date

# ------------------------
# Header normalization
# ------------------------
def _normalize_headers(cols):
    out = []
    for c in cols:
        s = str(c)
        s = s.replace("\n", " ")
        s = re.sub(r"\s+", "_", s)
        s = re.sub(r"\.", "", s)
        s = s.strip("_").upper()
        out.append(s)
    return out

# ------------------------
# Process single bill file -> DataFrame (formatted)
# ------------------------
def process_single_bill(input_path):
    try:
        raw_df = pd.read_excel(input_path, header=None, engine="openpyxl")
    except Exception as e:
        print(f"⚠️ Failed to read {input_path}: {e}")
        return pd.DataFrame()

    bill_num, bill_date = _extract_bill_meta(raw_df)

    # find the header row for the table
    header_row_idx = None
    header_pattern = re.compile(r"supply\s*inv(?:oice)?\.?\s*date", re.IGNORECASE)
    for i in range(raw_df.shape[0]):
        row_vals = ["" if v is None else str(v) for v in list(raw_df.iloc[i, :].values)]
        joined = " | ".join(row_vals)
        if header_pattern.search(joined):
            header_row_idx = i
            break
    if header_row_idx is None:
        # Could not find a table in this file
        print(f"⚠️ Header row not found in {os.path.basename(input_path)}. Skipping file.")
        return pd.DataFrame()

    # read the table with header
    try:
        df = pd.read_excel(input_path, header=header_row_idx, engine="openpyxl")
    except Exception as e:
        print(f"⚠️ Failed to parse table in {input_path}: {e}")
        return pd.DataFrame()

    df.columns = _normalize_headers(df.columns)

    # column mapping (flexible)
    col_map = {
        "ChallanNo":         ["SUPPLY_INVNO", "SUPPLY_INV_NO", "INV_NO", "INVOICE_NO"],
        "Destination":       ["DESTINATION", "CONSIGNEE_NAME"],
        "VehicleNo":         ["VEHICLE_REGNO", "VEHICLE_REG_NO", "VEHICLE_NO", "TRUCK_NO"],
        "DispatchDate":      ["SUPPLY_INV_DATE", "SUPPLY_INVOICE_DATE", "DATE", "DISPATCH_DATE"],
        "Dispatch Quantity": ["QTY", "QTY_MT", "QTYMT", "QUANTITY"],
        "Rate":              ["RATE/MT", "RATE_MT", "RATE_PER_MT", "RATE"],
        "Total Price":       ["TOTAL_VALUE", "TOTAL", "VALUE", "AMOUNT"],
        "LR":                ["LR_NO", "LRNUMBER", "LR"],
        "Deliver Num":       ["DELIVERY_NO", "DELIVERY_NUM", "DEL_NO"],
    }

    formatted_df = pd.DataFrame()
    for out_col, candidates in col_map.items():
        found = None
        for cand in candidates:
            if cand in df.columns:
                found = df[cand]
                break
        if found is None:
            formatted_df[out_col] = ""
        else:
            formatted_df[out_col] = found

    # filter out summary/blank rows: prefer to keep only rows where ChallanNo looks numeric
    if "ChallanNo" in formatted_df.columns:
        try:
            numeric_mask = pd.to_numeric(formatted_df["ChallanNo"], errors="coerce").notna()
            # if any row has numeric challan, keep only numeric challans (removes Total rows)
            if numeric_mask.any():
                formatted_df = formatted_df[numeric_mask].copy()
        except Exception:
            pass

    # attach Bill metadata and source filename
    formatted_df["Bill Num"] = bill_num if _is_meaningful_text(bill_num) else ""
    formatted_df["BILL DATE"] = bill_date if _is_meaningful_text(bill_date) else ""
    # do NOT add extra blank rows; return only rows of table

    return formatted_df

# ------------------------
# Convert numeric-like columns to numbers where possible (keep blanks as "")
# ------------------------
def _force_numeric_column(df, col_name):
    if col_name not in df.columns:
        return
    # convert to string, remove commas and spaces, attempt numeric conversion
    s = df[col_name].astype(str).fillna("").apply(lambda x: x.replace(",", "").strip())
    nums = pd.to_numeric(s, errors="coerce")
    # replace with floats where parseable, else empty string
    df[col_name] = [ (float(v) if not pd.isna(v) else "") for v in nums ]

# ------------------------
# Consolidate multiple files
# ------------------------
def convert_multiple_bills(folder, output_path):
    files = [f for f in os.listdir(folder) if f.lower().endswith((".xlsx", ".xls"))]
    if not files:
        print("❌ No Excel files found in folder:", folder)
        return

    all_parts = []
    for f in sorted(files):
        input_path = os.path.join(folder, f)
        print("Processing:", f)
        part_df = process_single_bill(input_path)
        if part_df is None or part_df.empty:
            print("  skipped (no table extracted).")
            continue
        # ensure same column order & names
        all_parts.append(part_df)

    if not all_parts:
        print("❌ No valid data extracted from any files.")
        return

    final_df = pd.concat(all_parts, ignore_index=True, sort=False)

    # ensure required columns exist in final_df
    required_columns = [
        "ChallanNo", "Destination", "VehicleNo", "DispatchDate",
        "Dispatch Quantity", "Rate", "Total Price", "LR", "Deliver Num",
        "Bill Num", "BILL DATE"
    ]
    final_df = final_df.reindex(columns=required_columns).fillna("")

    # Convert numeric-like columns to actual numbers where possible so Excel SUM works.
    numeric_columns = ["Dispatch Quantity", "Rate", "Total Price", "ChallanNo"]
    for col in numeric_columns:
        _force_numeric_column(final_df, col)

    # Write consolidated sheet (no blank rows between files)
    wb = Workbook()
    ws = wb.active
    ws.title = "JSW"

    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # header row
    ws.append(required_columns)
    for col_idx in range(1, len(required_columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border

    # data rows (write values preserving numeric types for numbers)
    for row in final_df.itertuples(index=False, name=None):
        ws.append(list(row))

    # style data cells
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(required_columns)):
        for cell in row_cells:
            cell.alignment = alignment
            cell.border = thin_border

    # autosize columns
    for col in ws.columns:
        max_width = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_width:
                max_width = len(val)
        ws.column_dimensions[col_letter].width = max_width + 2

    wb.save(output_path)
    print("✅ Consolidated output saved to:", output_path)
    print("Rows in consolidated sheet:", final_df.shape[0])

# ------------------------
# Entry point
# ------------------------
if __name__ == "__main__":
    # Fixed path for your device
    folder = r"H:\JswBillFormat"
    output_file = os.path.join(folder, "JSW_BILL_FORMATTED.xlsx")
    convert_multiple_bills(folder, output_file)
